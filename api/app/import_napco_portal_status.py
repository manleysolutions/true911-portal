"""NAPCO StarLink portal status import (dry-run-first).

Manley supplies the NAPCO StarLink fire communicators, pays the NAPCO MRC, and
rebills the subscriber; Manley manages them in the NAPCO portal. There is no
public API, but the portal exports an XLS/XLSX (or CSV) with device status, last
communication, name/address, trouble condition, etc. This command imports that
export and persists telemetry onto the matching devices.

It is the out-of-band monitoring path for ``classify().monitoring_source ==
'napco_xls_import'`` devices.

Safety:
  * DRY_RUN defaults TRUE — nothing written unless DRY_RUN=false.
  * Matches by serial number first, device_id second; name/address only flags a
    row REVIEW-REQUIRED (never auto-applied).
  * Staleness guard — a stale "last communication" never moves a device's
    heartbeat backwards, and a stale row does not regress network_status.
  * Writes only a whitelist of telemetry fields; never E911, lifecycle status,
    or any Assurance label; never fabricates a heartbeat.
  * Tenant-scoped (default restoration-hardware) — never touches another tenant.

Run:
    NAPCO_IMPORT_FILE=/path/to/export.xlsx python -m app.import_napco_portal_status
    DRY_RUN=false NAPCO_IMPORT_FILE=/path/to/export.xlsx python -m app.import_napco_portal_status
"""

from __future__ import annotations

import asyncio
import datetime as _dt
import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

NAPCO_TENANT = os.environ.get("NAPCO_IMPORT_TENANT", "restoration-hardware")

# Only these Device columns may be written here — defence in depth.
ALLOWED_DEVICE_FIELDS = frozenset({
    "last_heartbeat", "last_network_event", "network_status",
    "telemetry_source", "carrier", "firmware_version",
})
STALENESS_GUARDED = ("last_heartbeat", "last_network_event")

# Canonical field -> ordered header substrings to look for (case-insensitive).
_COLUMN_CANDIDATES = {
    "serial": ("serial", "esn", "device serial"),
    "device_id": ("device id", "device_id", "deviceid"),
    "portal_status": ("comm status", "communication status", "status"),
    "last_comm": ("last communication", "last comm", "last check", "last signal",
                  "last report", "last seen", "last received"),
    "name": ("account name", "site name", "name"),
    "address": ("address", "location"),
    "trouble": ("trouble", "fault", "alarm condition"),
    "carrier": ("carrier", "network"),
    "model": ("model", "device type", "communicator", "type"),
    "config": ("configuration", "config", "profile"),
}

_HEALTHY_STATUS = {"online", "normal", "ok", "good", "communicating", "ready", "active", "clear"}
_OFFLINE_STATUS = {"offline", "no comm", "no communication", "nocomm", "fault",
                   "fail", "failed", "lost", "disconnected", "down"}
_TROUBLE_CLEAR = {"", "none", "no", "normal", "ok", "clear", "0", "false", "n/a"}


# ── pure helpers (unit-tested) ───────────────────────────────────────────
def _norm(v) -> str:
    return v.strip() if isinstance(v, str) else ("" if v is None else str(v).strip())


def build_column_map(headers: list) -> dict:
    """Map canonical field -> column index from a header row. First match wins;
    a more specific candidate (e.g. 'comm status') is tried before 'status'."""
    lowered = [(_norm(h).lower()) for h in headers]
    used: set[int] = set()
    out: dict[str, int] = {}
    for field, candidates in _COLUMN_CANDIDATES.items():
        for cand in candidates:
            idx = next((i for i, h in enumerate(lowered) if cand in h and i not in used), None)
            if idx is not None:
                out[field] = idx
                used.add(idx)
                break
    return out


def parse_last_comm(value) -> _dt.datetime | None:
    """Robustly parse a NAPCO 'last communication' cell to a tz-aware UTC datetime."""
    if value is None or value == "":
        return None
    if isinstance(value, _dt.datetime):
        return value if value.tzinfo else value.replace(tzinfo=_dt.timezone.utc)
    if isinstance(value, _dt.date):
        return _dt.datetime(value.year, value.month, value.day, tzinfo=_dt.timezone.utc)
    if isinstance(value, (int, float)):
        # Excel serial date (days since 1899-12-30).
        try:
            return (_dt.datetime(1899, 12, 30, tzinfo=_dt.timezone.utc)
                    + _dt.timedelta(days=float(value)))
        except (ValueError, OverflowError):
            return None
    s = _norm(value)
    for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d",
                "%m/%d/%Y %I:%M:%S %p", "%m/%d/%Y %H:%M:%S", "%m/%d/%Y %H:%M",
                "%m/%d/%Y %I:%M %p", "%m/%d/%Y", "%m/%d/%y %H:%M", "%m/%d/%y"):
        try:
            return _dt.datetime.strptime(s, fmt).replace(tzinfo=_dt.timezone.utc)
        except ValueError:
            continue
    try:
        dt = _dt.datetime.fromisoformat(s)
        return dt if dt.tzinfo else dt.replace(tzinfo=_dt.timezone.utc)
    except ValueError:
        return None


def map_status_to_network(portal_status, trouble) -> str:
    """NAPCO portal status + trouble -> True911 network_status. Trouble/offline
    are NON-healthy."""
    t = _norm(trouble).lower()
    if t and t not in _TROUBLE_CLEAR:
        return "trouble"
    s = _norm(portal_status).lower()
    if s in _HEALTHY_STATUS:
        return "online"
    if s in _OFFLINE_STATUS:
        return "offline"
    return "unknown"


def parse_napco_row(row: list, column_map: dict) -> dict:
    """Extract canonical fields from one spreadsheet row (list of cell values)."""
    def cell(field):
        idx = column_map.get(field)
        return row[idx] if idx is not None and idx < len(row) else None

    return {
        "serial": _norm(cell("serial")),
        "device_id": _norm(cell("device_id")),
        "portal_status": _norm(cell("portal_status")),
        "last_comm": parse_last_comm(cell("last_comm")),
        "name": _norm(cell("name")),
        "address": _norm(cell("address")),
        "trouble": _norm(cell("trouble")),
        "carrier": _norm(cell("carrier")),
        "model": _norm(cell("model")),
        "config": _norm(cell("config")),
        "network_status": map_status_to_network(cell("portal_status"), cell("trouble")),
    }


def match_device(parsed: dict, by_serial: dict, by_device_id: dict) -> tuple:
    """Return (device_or_None, method, review_required). Serial first, device_id
    second; otherwise name/address fallback flagged review-required."""
    serial = parsed.get("serial", "")
    if serial and serial.lower() in by_serial:
        return by_serial[serial.lower()], "serial", False
    did = parsed.get("device_id", "")
    if did and did in by_device_id:
        return by_device_id[did], "device_id", False
    return None, "name_address_fallback", True


def _coerce_utc(dt):
    if dt is None or not isinstance(dt, _dt.datetime):
        return dt
    return dt if dt.tzinfo is not None else dt.replace(tzinfo=_dt.timezone.utc)


def _is_newer(proposed, current) -> bool:
    if current is None:
        return True
    if proposed is None:
        return False
    try:
        return _coerce_utc(proposed) > _coerce_utc(current)
    except TypeError:
        return True


def compute_napco_updates(parsed: dict, current: dict, *, now=None) -> tuple:
    """Pure: telemetry field updates for a matched device, with staleness guard.
    Returns (kept_updates, notes). Never writes E911/status/Assurance."""
    notes: list[str] = []
    dev: dict = {"telemetry_source": "napco_portal"}  # mark the monitoring source

    last_comm = parsed.get("last_comm")
    fresh = last_comm is not None and _is_newer(last_comm, current.get("last_heartbeat"))
    if last_comm is None:
        notes.append("no parseable last communication — no heartbeat/status applied")
    elif fresh:
        dev["last_heartbeat"] = last_comm
        dev["last_network_event"] = last_comm
        dev["network_status"] = parsed.get("network_status") or "unknown"
    else:
        notes.append(f"skipped stale last communication ({last_comm} <= stored "
                     f"{current.get('last_heartbeat')}) — no regression")

    if parsed.get("carrier") and not _norm(current.get("carrier")):
        dev["carrier"] = parsed["carrier"]

    # Defence in depth: drop anything not whitelisted.
    kept = {k: v for k, v in dev.items() if k in ALLOWED_DEVICE_FIELDS}
    for k in dev:
        if k not in ALLOWED_DEVICE_FIELDS:
            notes.append(f"ignored non-telemetry field {k!r}")
    return kept, notes


def safe_metadata(parsed: dict) -> dict:
    """Device-identity/diagnostic fields safe to archive — no credentials/secrets."""
    return {k: (v.isoformat() if isinstance(v, _dt.datetime) else v)
            for k, v in parsed.items()}


# ── file reading ─────────────────────────────────────────────────────────
def read_rows(path: str) -> tuple:
    """Return (headers, data_rows) from an .xlsx/.xls or .csv export."""
    lower = path.lower()
    if lower.endswith((".xlsx", ".xlsm", ".xls")):
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        wb.close()
    elif lower.endswith(".csv"):
        import csv
        with open(path, "r", encoding="utf-8-sig", newline="") as fh:
            rows = [list(r) for r in csv.reader(fh)]
    else:
        raise ValueError(f"unsupported file type: {path} (use .xlsx/.xls/.csv)")
    rows = [r for r in rows if any(_norm(c) for c in r)]  # drop blank rows
    if not rows:
        return [], []
    return rows[0], rows[1:]


# ── DB orchestration ─────────────────────────────────────────────────────
async def run(*, dry_run: bool = True, path: str | None = None, tenant_id: str = NAPCO_TENANT) -> dict:
    import uuid
    from sqlalchemy import select
    from app.database import AsyncSessionLocal
    from app.models.device import Device
    from app.models.integration_payload import IntegrationPayload
    from app.services.audit_logger import log_audit

    now = _dt.datetime.now(_dt.timezone.utc)
    summary = {"rows": 0, "matched_serial": 0, "matched_device_id": 0, "review_required": 0,
               "updated": 0, "stale_skipped": 0, "offline_or_trouble": 0, "notes": []}

    if not path:
        print("No NAPCO_IMPORT_FILE given. Set it to the portal XLS/XLSX/CSV export.")
        return summary
    try:
        headers, data_rows = read_rows(path)
    except (OSError, ValueError, ImportError) as exc:
        print(f"ERROR reading {path!r}: {exc}. Nothing written.")
        return summary
    column_map = build_column_map(headers)
    if "serial" not in column_map and "device_id" not in column_map:
        print(f"ERROR: could not find a serial or device-id column in headers {headers}. "
              "Nothing written.")
        return summary
    print(f"  columns detected: { {k: headers[v] for k, v in column_map.items()} }")

    async with AsyncSessionLocal() as db:
        devices = (await db.execute(
            select(Device).where(Device.tenant_id == tenant_id))).scalars().all()
        by_serial = {_norm(d.serial_number).lower(): d for d in devices if _norm(d.serial_number)}
        by_device_id = {d.device_id: d for d in devices}

        for raw in data_rows:
            parsed = parse_napco_row(raw, column_map)
            summary["rows"] += 1
            if parsed["network_status"] in ("offline", "trouble"):
                summary["offline_or_trouble"] += 1

            device, method, review = match_device(parsed, by_serial, by_device_id)
            if review or device is None:
                summary["review_required"] += 1
                print(f"  REVIEW  serial={parsed['serial']!r} name={parsed['name']!r} "
                      f"addr={parsed['address']!r} — no serial/device_id match; operator review required")
                continue
            summary["matched_serial" if method == "serial" else "matched_device_id"] += 1

            current = {"last_heartbeat": device.last_heartbeat, "carrier": device.carrier}
            kept, notes = compute_napco_updates(parsed, current, now=now)
            for n in notes:
                if n.startswith("skipped stale"):
                    summary["stale_skipped"] += 1
                print(f"      · {n}")
            print(f"  {device.device_id:20} match={method} status={parsed['portal_status']!r} "
                  f"-> net={kept.get('network_status', '(unchanged)')} "
                  f"last_comm={parsed['last_comm']}")

            if dry_run:
                print(f"      would update {kept}")
                continue

            for f, v in kept.items():
                setattr(device, f, v)
            db.add(IntegrationPayload(
                payload_id=f"napco-{uuid.uuid4().hex[:12]}",
                source="napco_portal", direction="inbound",
                body=safe_metadata(parsed), processed=True))
            summary["updated"] += 1
            await log_audit(
                db, tenant_id, "device_health", "napco_xls_import",
                f"NAPCO portal status import for {device.device_id} (match={method})",
                actor="import_napco_portal_status", target_type="device",
                target_id=device.device_id, site_id=device.site_id, device_id=device.device_id,
                detail={"changes": {k: str(v) for k, v in kept.items()},
                        "portal_status": parsed["portal_status"], "trouble": parsed["trouble"],
                        "match_method": method, "stale_notes": notes})

        if dry_run:
            await db.rollback()
            print("\nDRY RUN — no changes committed.")
        else:
            await db.commit()
            print("\nCommitted.")
    return summary


def main() -> None:
    dry_run = os.environ.get("DRY_RUN", "true").strip().lower() not in ("0", "false", "no", "off")
    path = os.environ.get("NAPCO_IMPORT_FILE") or None
    print("=" * 66)
    print(f"NAPCO StarLink portal status import — tenant '{NAPCO_TENANT}'")
    print(f"  mode: {'DRY RUN (no writes)' if dry_run else 'APPLY (telemetry fields only)'}")
    print(f"  file: {path or '(none)'}")
    print("=" * 66)
    try:
        summary = asyncio.run(run(dry_run=dry_run, path=path))
    except Exception as exc:  # pragma: no cover - connectivity edge
        print(f"\nERROR: NAPCO import aborted — {type(exc).__name__}: {exc}")
        raise SystemExit(1)

    print("\n" + "=" * 66)
    print("SUMMARY")
    print("=" * 66)
    for k in ("rows", "matched_serial", "matched_device_id", "review_required",
              "updated", "stale_skipped", "offline_or_trouble"):
        print(f"  {k:20}: {summary[k]}")


if __name__ == "__main__":
    main()
